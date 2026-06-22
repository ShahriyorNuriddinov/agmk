import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook

from app.db.session import get_db
from app.models.organization import Department, Section
from app.models.worker import Worker, WorkerStatus
from app.models.briefing import Briefing, BriefingStatus
from app.models.knowledge_check import KnowledgeCheck, CheckStatus
from app.models.user import User
from app.core.deps import get_current_user

router = APIRouter()


@router.get("/briefings/excel")
async def export_briefings_excel(
    department_id: int = Query(None),
    status: BriefingStatus = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Briefing).join(Worker)
    if department_id:
        query = query.where(Briefing.department_id == department_id)
    if status:
        query = query.where(Briefing.status == status)
    query = query.order_by(Briefing.conducted_at.desc())
    result = await db.execute(query)
    briefings = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Инструктажи"
    ws.append([
        "№", "Табельный номер", "Работник", "Вид инструктажа",
        "Дата проведения", "Инструктирующий", "Следующая дата", "Статус",
        "Подтвержден работником", "Подтвержден инструктором"
    ])

    for i, b in enumerate(briefings, 1):
        worker_r = await db.execute(select(Worker).where(Worker.id == b.worker_id))
        worker = worker_r.scalar_one_or_none()
        ws.append([
            i,
            worker.personnel_number if worker else "",
            f"{worker.last_name} {worker.first_name}" if worker else "",
            b.briefing_type.value,
            b.conducted_at.strftime("%d.%m.%Y %H:%M"),
            b.instructor_name,
            b.next_briefing_date.strftime("%d.%m.%Y") if b.next_briefing_date else "-",
            b.status.value,
            "Да" if b.worker_confirmed else "Нет",
            "Да" if b.instructor_confirmed else "Нет",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=briefings_{date.today()}.xlsx"},
    )


@router.get("/knowledge-checks/excel")
async def export_knowledge_checks_excel(
    status: CheckStatus = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(KnowledgeCheck).join(Worker)
    if status:
        query = query.where(KnowledgeCheck.status == status)
    query = query.order_by(KnowledgeCheck.check_date.desc())
    result = await db.execute(query)
    checks = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка знаний"
    ws.append([
        "№", "Табельный номер", "Работник", "Дата проверки",
        "№ протокола", "Результат", "Следующая дата", "Статус",
        "Подтвержден работником", "Подтвержден председателем"
    ])

    for i, c in enumerate(checks, 1):
        worker_r = await db.execute(select(Worker).where(Worker.id == c.worker_id))
        worker = worker_r.scalar_one_or_none()
        ws.append([
            i,
            worker.personnel_number if worker else "",
            f"{worker.last_name} {worker.first_name}" if worker else "",
            c.check_date.strftime("%d.%m.%Y"),
            c.protocol_number,
            c.result.value,
            c.next_check_date.strftime("%d.%m.%Y"),
            c.status.value,
            "Да" if c.worker_confirmed else "Нет",
            "Да" if c.chairman_confirmed else "Нет",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=knowledge_checks_{date.today()}.xlsx"},
    )


@router.get("/overdue-workers/excel")
async def export_overdue_workers(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Workers with overdue briefings
    overdue_b_r = await db.execute(
        select(Briefing).where(Briefing.status == BriefingStatus.OVERDUE)
    )
    overdue_briefings = overdue_b_r.scalars().all()

    # Workers with overdue checks
    overdue_c_r = await db.execute(
        select(KnowledgeCheck).where(KnowledgeCheck.status == CheckStatus.OVERDUE)
    )
    overdue_checks = overdue_c_r.scalars().all()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Просроченные инструктажи"
    ws1.append(["№", "Табельный номер", "ФИО", "Вид инструктажа", "Дата просрочки"])

    for i, b in enumerate(overdue_briefings, 1):
        worker_r = await db.execute(select(Worker).where(Worker.id == b.worker_id))
        worker = worker_r.scalar_one_or_none()
        ws1.append([
            i,
            worker.personnel_number if worker else "",
            f"{worker.last_name} {worker.first_name}" if worker else "",
            b.briefing_type.value,
            b.next_briefing_date.strftime("%d.%m.%Y") if b.next_briefing_date else "-",
        ])

    ws2 = wb.create_sheet("Просроченные проверки знаний")
    ws2.append(["№", "Табельный номер", "ФИО", "Дата следующей проверки"])

    for i, c in enumerate(overdue_checks, 1):
        worker_r = await db.execute(select(Worker).where(Worker.id == c.worker_id))
        worker = worker_r.scalar_one_or_none()
        ws2.append([
            i,
            worker.personnel_number if worker else "",
            f"{worker.last_name} {worker.first_name}" if worker else "",
            c.next_check_date.strftime("%d.%m.%Y"),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=overdue_workers_{date.today()}.xlsx"},
    )
